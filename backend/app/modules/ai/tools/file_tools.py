"""
文件解析工具

录单员核心能力：解析 Excel / CSV，预览前 N 行，便于 LLM 理解结构后做字段映射。
PDF / 图片 OCR 远期接入。
"""

from __future__ import annotations

from pathlib import Path
from typing import Optional

from pydantic import BaseModel, Field

from app.modules.ai.tools.base import ToolContext, ToolResult
from app.modules.ai.tools.registry import register_tool

# 与 ai 文件上传接口落盘根目录保持一致
AI_UPLOAD_ROOT = Path(__file__).resolve().parents[3] / "uploads" / "ai_attach"


class ParseExcelParams(BaseModel):
    file_id: str = Field(..., description="AI 附件文件ID（来自 /ai/file/upload 返回）")
    sheet_name: Optional[str] = Field(
        None, description="工作表名，不传则取第一个 sheet"
    )
    max_preview_rows: int = Field(
        20, ge=1, le=200, description="预览行数，默认20"
    )


def _resolve_safe_path(file_id: str) -> Path:
    """防穿越：file_id 不允许含路径分隔符"""
    if "/" in file_id or "\\" in file_id or ".." in file_id:
        raise ValueError("非法 file_id")
    p = AI_UPLOAD_ROOT / file_id
    if not p.exists() or not p.is_file():
        raise FileNotFoundError(f"文件不存在: {file_id}")
    return p


@register_tool(
    code="file.parse_excel",
    name="解析 Excel/CSV",
    category="file",
    description=(
        "解析用户上传的 Excel(.xlsx/.xls) 或 CSV 文件，返回表头、列类型推断与样本数据。"
        "用于录单员场景下的字段映射前置分析。"
    ),
    params_schema=ParseExcelParams,
    risk_level="low",
)
async def parse_excel(ctx: ToolContext, **kwargs) -> ToolResult:
    params = ParseExcelParams(**kwargs)
    try:
        path = _resolve_safe_path(params.file_id)
    except (ValueError, FileNotFoundError) as e:
        return ToolResult(success=False, error=str(e))

    suffix = path.suffix.lower()
    try:
        if suffix == ".csv":
            preview = _parse_csv(path, params.max_preview_rows)
        elif suffix in (".xlsx", ".xls"):
            preview = _parse_excel(path, params.sheet_name, params.max_preview_rows)
        else:
            return ToolResult(
                success=False,
                error=f"暂不支持的文件类型 {suffix}（支持 .xlsx/.xls/.csv）",
            )
    except Exception as e:  # noqa: BLE001
        return ToolResult(success=False, error=f"解析失败: {e}")

    return ToolResult(
        success=True,
        data=preview,
        message=(
            f"解析成功，共 {preview.get('total_rows', 0)} 行，"
            f"返回前 {len(preview.get('rows', []))} 行预览"
        ),
    )


def _parse_csv(path: Path, max_rows: int) -> dict:
    import csv

    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        all_rows = list(reader)
    if not all_rows:
        return {"headers": [], "rows": [], "total_rows": 0}
    headers = [str(h).strip() for h in all_rows[0]]
    body = all_rows[1:]
    rows = [
        {headers[i]: (cell if i < len(row) else "") for i, cell in enumerate(row)}
        for row in body[:max_rows]
    ]
    return {
        "format": "csv",
        "headers": headers,
        "rows": rows,
        "total_rows": len(body),
    }


def _parse_excel(path: Path, sheet_name: Optional[str], max_rows: int) -> dict:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError(
            "未安装 openpyxl，请执行 `pip install openpyxl>=3.1.5`"
        ) from e

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
        sheet_names = wb.sheetnames

        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return {
                "format": "excel",
                "sheet_names": sheet_names,
                "current_sheet": ws.title,
                "headers": [],
                "rows": [],
                "total_rows": 0,
            }
        headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(header_row)]

        body, total = [], 0
        for row in rows_iter:
            total += 1
            if len(body) < max_rows:
                body.append(
                    {
                        headers[i]: (cell if i < len(row) else None)
                        for i, cell in enumerate(row)
                    }
                )
        return {
            "format": "excel",
            "sheet_names": sheet_names,
            "current_sheet": ws.title,
            "headers": headers,
            "rows": body,
            "total_rows": total,
        }
    finally:
        wb.close()


class MapColumnsParams(BaseModel):
    file_id: str = Field(..., description="AI 附件文件ID")
    column_mapping: dict[str, str] = Field(
        ...,
        description=(
            "源列名 -> 目标字段名 的映射；目标字段名应与目标 Service 的字段对齐"
        ),
    )
    sheet_name: Optional[str] = Field(None, description="工作表名")
    max_rows: int = Field(500, ge=1, le=2000, description="本次最多映射行数")


@register_tool(
    code="file.map_columns",
    name="按列映射转换为目标行",
    category="file",
    description=(
        "依据 column_mapping 把 Excel/CSV 数据转换为目标字段命名的行列表。"
        "结果可直接传给 waybill.batch_create 等批量入库工具。"
    ),
    params_schema=MapColumnsParams,
    risk_level="low",
)
async def map_columns(ctx: ToolContext, **kwargs) -> ToolResult:
    params = MapColumnsParams(**kwargs)
    try:
        path = _resolve_safe_path(params.file_id)
    except (ValueError, FileNotFoundError) as e:
        return ToolResult(success=False, error=str(e))

    suffix = path.suffix.lower()
    try:
        if suffix == ".csv":
            preview = _parse_csv(path, params.max_rows)
        elif suffix in (".xlsx", ".xls"):
            preview = _parse_excel(path, params.sheet_name, params.max_rows)
        else:
            return ToolResult(success=False, error=f"暂不支持 {suffix}")
    except Exception as e:  # noqa: BLE001
        return ToolResult(success=False, error=f"读取失败: {e}")

    headers = preview.get("headers", [])
    src_rows = preview.get("rows", [])
    mapping = params.column_mapping
    unknown_src = [c for c in mapping.keys() if c not in headers]

    target_rows = []
    for row in src_rows:
        out: dict = {}
        for src_col, dst_field in mapping.items():
            v = row.get(src_col)
            if v is not None and v != "":
                out[dst_field] = v
        if out:
            target_rows.append(out)

    return ToolResult(
        success=True,
        data={
            "rows": target_rows,
            "row_count": len(target_rows),
            "unknown_source_columns": unknown_src,
        },
        message=f"已映射 {len(target_rows)} 行；未识别的源列 {len(unknown_src)} 个",
    )
