"""Excel 账单导入连接器"""

from __future__ import annotations

from io import BytesIO
from typing import List

from app.modules.client.services.energy.connectors.registry import (
    ConnectorContext,
    RawRecord,
    register_connector,
)
from app.modules.client.services.energy.normalizer import json_safe_value


@register_connector(code="excel", name="Excel 账单导入", sync_modes=["manual"],
                    description="按模板上传供应商账单，解析为原始记录")
class ExcelConnector:
    TEMPLATE_HEADERS = [
        "流水号", "卡号", "车牌号", "站点", "能源类型", "商品",
        "数量", "单位", "单价", "金额", "消费时间", "里程",
    ]

    async def fetch(self, ctx: ConnectorContext) -> List[RawRecord]:
        rows = (ctx.extra or {}).get("rows") or []
        return [RawRecord(data=r, external_transaction_id=r.get("流水号") or r.get("externalTransactionId"))
                for r in rows]

    @staticmethod
    def parse_workbook(content: bytes) -> List[dict]:
        from openpyxl import load_workbook

        wb = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
        except StopIteration:
            return []
        out: List[dict] = []
        for row in rows_iter:
            item = {}
            empty = True
            for i, h in enumerate(headers):
                if not h:
                    continue
                val = row[i] if i < len(row) else None
                if val not in (None, ""):
                    empty = False
                item[h] = json_safe_value(val)
            if not empty:
                out.append(item)
        return out
