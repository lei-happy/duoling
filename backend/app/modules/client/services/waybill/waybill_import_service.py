"""
计划批量导入服务（Phase 6）

工作流：
  1) 接口接收 Excel 文件 → 解析 → 写 biz_waybill_import_batch + biz_waybill_import_row(raw_data_json)
  2) 同步遍历每行：标准化 + 字段校验 → 创建 Waybill / WaybillCargo
  3) 创建后由 WaybillService.create_waybill 内部 enqueue freight_calc_task
  4) 行级失败写 validate_status='failed' + validate_message；批次更新统计

Excel 列约定（按表头中文匹配，大小写不敏感、去空格）：
  - 计划编号 (可空，自动生成)
  - 客户名称（必填；后台按名称匹配 biz_customer.id，勿填客户ID）
  - 出发地 / 目的地（必填其一或组合；逐级用 / 分隔，如「广东省/广州市/天河区」；
    后台用 StandardizeService 解析区划编码与 region_id，勿填编码/区域ID）
  - 商品车品牌（或兼容旧表头「货物品牌」）/ 车型 / VIN码（新模板）或 数量（旧模板）
  - 计划开单时间 / 要求装车时间 / 要求送达时间 (yyyy-mm-dd HH:MM)
  - 经销商名称 / 联系人 / 电话 / 地址
  - 备注

多明细：以「+」串接同一行的多车型，例如：
  - 商品车品牌列 = "比亚迪+长安"
  - 车型列       = "汉EV+逸动"
  - VIN码列      = "LGWHXXXXXXXXXXXX1+LGWHYYYYYYYYYYYY2"（分段数须与品牌/车型一致）

旧模板仍支持「数量」列：后台按台数拆成多行，并为每行生成占位 VIN（ZIMP…）以满足创建校验。

兼容旧模板：仍识别 客户ID、出发地/目的地编码与区域ID、货物品牌 等列。
"""

from __future__ import annotations

import io
from typing import Any, Optional

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.common.exceptions import BizException
from app.modules.client.models.partner.customer import Customer
from app.modules.client.models.waybill.waybill_import import (
    WaybillImportBatch,
    WaybillImportRow,
)
from app.modules.client.schemas.waybill.waybill import (
    WaybillCargoLineIn,
    WaybillCreate,
    normalize_waybill_vin,
)
from app.modules.client.services.waybill.waybill_service import WaybillService


# Excel 表头映射（中文 → 标准 key）
HEADER_MAP: dict[str, str] = {
    "计划编号": "waybillNo",
    "运单编号": "waybillNo",  # 兼容旧 Excel 模板
    "客户名称": "customerName",
    "客户id": "customerId",
    "客户ID": "customerId",
    "出发地": "origin",
    "目的地": "destination",
    "出发地编码": "originCode",
    "目的地编码": "destinationCode",
    "出发地区域id": "originRegionId",
    "目的地区域id": "destinationRegionId",
    "货物品牌": "vehicleBrand",
    "商品车品牌": "vehicleBrand",
    "品牌": "vehicleBrand",
    "车型": "vehicleModel",
    "vin码": "vin",
    "车架号": "vin",
    "vin": "vin",
    "数量": "quantity",
    "计划开单时间": "planIssueTime",
    "要求装车时间": "requiredLoadTime",
    "要求送达时间": "requiredDeliverTime",
    "经销商名称": "dealerName",
    "经销商联系人": "dealerContact",
    "经销商电话": "dealerPhone",
    "经销商地址": "dealerAddress",
    "备注": "remark",
}

# 下载模板用表头顺序（须与 HEADER_MAP 中文键一致，便于解析首行识别列）
IMPORT_TEMPLATE_HEADERS: tuple[str, ...] = (
    "计划编号",
    "客户名称",
    "出发地",
    "目的地",
    "商品车品牌",
    "车型",
    "VIN码",
    "计划开单时间",
    "要求装车时间",
    "要求送达时间",
    "经销商名称",
    "经销商联系人",
    "经销商电话",
    "经销商地址",
    "备注",
)


def _norm_header(s: Any) -> str:
    if s is None:
        return ""
    return str(s).strip().replace("\u3000", "").replace(" ", "").lower()


def _split_multi(val: Any) -> list[str]:
    if val is None:
        return []
    return [s.strip() for s in str(val).split("+") if s.strip()]


def _parse_int(val: Any, default: Optional[int] = None) -> Optional[int]:
    if val is None or val == "":
        return default
    try:
        return int(float(val))
    except (TypeError, ValueError):
        return default


def _norm_region_code(val: Any) -> Optional[str]:
    """Excel 中区划码可能是数字类型或带小数点的字符串，统一为整数字符串。"""
    if val is None or val == "":
        return None
    if isinstance(val, bool):
        return None
    if isinstance(val, (int, float)):
        try:
            return str(int(val))
        except (OverflowError, ValueError):
            return str(val).strip() or None
    s = str(val).strip()
    if not s:
        return None
    if s.isdigit():
        return s
    try:
        f = float(s)
        if f > 0 and f == int(f):
            return str(int(f))
    except (TypeError, ValueError):
        pass
    return s


def _parse_dt(val: Any):
    if val is None or val == "":
        return None
    if hasattr(val, "isoformat"):
        return val
    s = str(val).strip()
    from datetime import datetime as _dt
    fmts = (
        "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M",
        "%Y-%m-%d", "%Y/%m/%d",
    )
    for f in fmts:
        try:
            return _dt.strptime(s, f)
        except ValueError:
            continue
    return None


def _str_cell(row: dict, key: str) -> Optional[str]:
    v = row.get(key)
    if v is None or v == "":
        return None
    s = str(v).strip()
    return s if s else None


def _waybill_no_from_row(row: dict) -> Optional[str]:
    return _str_cell(row, "waybillNo")


class WaybillImportService:

    @staticmethod
    async def _resolve_customer_for_import(
        db: AsyncSession, row: dict,
    ) -> tuple[Optional[int], Optional[str]]:
        """优先按客户名称精确匹配启用客户；无名称时可退回旧模板的客户ID。"""
        name = _str_cell(row, "customerName")
        if name:
            r = await db.execute(
                select(Customer).where(
                    Customer.customer_name == name,
                    Customer.is_deleted == 0,
                    Customer.status == 1,
                ).order_by(Customer.id.asc())
            )
            rows = list(r.scalars().all())
            if len(rows) == 0:
                raise BizException(
                    f"未找到启用状态的客户「{name}」，请与客户档案中的名称完全一致"
                )
            if len(rows) > 1:
                raise BizException(
                    f"客户名称「{name}」对应多条启用记录，请在客户管理中区分后再导入"
                )
            c = rows[0]
            return c.id, (c.customer_name or name)

        cid = _parse_int(row.get("customerId"))
        if cid is not None:
            r = await db.execute(
                select(Customer).where(
                    Customer.id == cid,
                    Customer.is_deleted == 0,
                )
            )
            c = r.scalar_one_or_none()
            if c is None:
                raise BizException(f"客户ID「{cid}」不存在或已删除")
            if int(c.status or 0) != 1:
                raise BizException(f"客户ID「{cid}」已停用，无法导入")
            cn2 = _str_cell(row, "customerName")
            return cid, (cn2 or (c.customer_name or None))

        raise BizException("请填写客户名称")

    @staticmethod
    def _import_synthetic_vin(row_no: int, seq: int) -> str:
        """旧模板按台数拆行时的占位 VIN（10~50 位，仅导入兼容用）。"""
        return f"ZIMP{int(row_no) % 100000:05d}{int(seq) % 100000000:08d}"

    @staticmethod
    async def _build_waybill_create_from_row(
        db: AsyncSession, row: dict, *, row_no: int = 0,
    ) -> WaybillCreate:
        """解析行 → WaybillCreate（客户名称匹配 ID；地区仍走 create 内 hydrate）。"""
        customer_id, customer_name = (
            await WaybillImportService._resolve_customer_for_import(db, row)
        )

        cargoes: list[WaybillCargoLineIn] = []
        brands = _split_multi(row.get("vehicleBrand"))
        models = _split_multi(row.get("vehicleModel"))
        vins_raw = _split_multi(row.get("vin"))
        qtys = _split_multi(row.get("quantity"))

        use_vin_column = bool(vins_raw)
        if use_vin_column:
            if len(brands) != len(models) or len(brands) != len(vins_raw):
                raise BizException(
                    "使用 VIN 导入时，商品车品牌、车型、VIN码三列以「+」分段的数量须完全一致"
                )
            seq = 0
            for i in range(len(brands)):
                brand = brands[i]
                model = models[i]
                vn = vins_raw[i]
                nv = normalize_waybill_vin(vn)
                if not nv or len(nv) < 10 or len(nv) > 50:
                    raise BizException(
                        f"VIN码第 {i + 1} 段无效（须为 10~50 位字母或数字）"
                    )
                if not brand and not model:
                    continue
                cargoes.append(
                    WaybillCargoLineIn(
                        vehicleBrand=brand,
                        vehicleModel=model,
                        vin=nv,
                        quantity=1,
                        sortOrder=seq,
                    )
                )
                seq += 1
        else:
            n = max(len(brands), len(models), len(qtys), 1)
            seq = 0
            for i in range(n):
                brand = brands[i] if i < len(brands) else (
                    brands[0] if brands else None
                )
                model = models[i] if i < len(models) else (
                    models[0] if models else None
                )
                qty_raw = qtys[i] if i < len(qtys) else (qtys[0] if qtys else None)
                qty = _parse_int(qty_raw, 1) or 1
                if not brand and not model:
                    continue
                if qty < 1:
                    qty = 1
                for _k in range(int(qty)):
                    seq += 1
                    cargoes.append(
                        WaybillCargoLineIn(
                            vehicleBrand=brand,
                            vehicleModel=model,
                            vin=WaybillImportService._import_synthetic_vin(
                                row_no, seq,
                            ),
                            quantity=1,
                            sortOrder=len(cargoes),
                        )
                    )
        if not cargoes:
            raise BizException("缺少商品车品牌/车型及 VIN 或数量")

        origin = _str_cell(row, "origin")
        dest = _str_cell(row, "destination")
        if not origin or not dest:
            raise BizException("请同时填写出发地与目的地（多级行政区用 / 分隔）")

        return WaybillCreate(
            waybillNo=_waybill_no_from_row(row),
            customerId=customer_id,
            customerName=customer_name,
            origin=origin,
            originCode=_norm_region_code(row.get("originCode")),
            originRegionId=_parse_int(row.get("originRegionId")),
            destination=dest,
            destinationCode=_norm_region_code(row.get("destinationCode")),
            destinationRegionId=_parse_int(row.get("destinationRegionId")),
            cargoes=cargoes,
            planIssueTime=_parse_dt(row.get("planIssueTime")),
            requiredLoadTime=_parse_dt(row.get("requiredLoadTime")),
            requiredDeliverTime=_parse_dt(row.get("requiredDeliverTime")),
            dealerName=_str_cell(row, "dealerName"),
            dealerContact=_str_cell(row, "dealerContact"),
            dealerPhone=_str_cell(row, "dealerPhone"),
            dealerAddress=_str_cell(row, "dealerAddress"),
            remark=_str_cell(row, "remark"),
        )

    @staticmethod
    def build_template_workbook_bytes() -> bytes:
        """生成带标准表头的空 xlsx，供用户下载填写。"""
        try:
            import openpyxl  # type: ignore
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError as e:  # pragma: no cover
            raise BizException("缺少 openpyxl 依赖") from e

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "计划导入"
        headers = list(IMPORT_TEMPLATE_HEADERS)
        ws.append(headers)
        # 空行便于用户从第 3 行起填；解析时会跳过全空行
        ws.append([""] * len(headers))

        header_fill = PatternFill(
            "solid", fgColor="FFE7EEF8"
        )
        header_font = Font(bold=True, size=11)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        widths = (
            14, 18, 28, 28, 14, 14, 22, 18, 18, 18, 16, 12, 14, 24, 20,
        )
        for i, w in enumerate(widths, start=1):
            if i <= len(headers):
                ws.column_dimensions[get_column_letter(i)].width = w

        bio = io.BytesIO()
        wb.save(bio)
        return bio.getvalue()

    @staticmethod
    def _parse_excel(file_bytes: bytes) -> list[dict]:
        """解析 xlsx → 返回每行 dict（key 已映射到 schema 字段）"""
        try:
            import openpyxl  # type: ignore
        except ImportError as e:  # pragma: no cover
            raise BizException("缺少 openpyxl 依赖") from e

        wb = openpyxl.load_workbook(
            io.BytesIO(file_bytes), data_only=True, read_only=True
        )
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        header_row = rows[0]
        col_map: dict[int, str] = {}
        for idx, cell in enumerate(header_row):
            key = HEADER_MAP.get(_norm_header(cell), None)
            # 同时尝试原始表头（去空格）
            if not key:
                k2 = HEADER_MAP.get(str(cell).strip() if cell else "", None)
                if k2:
                    key = k2
            if key:
                col_map[idx] = key

        out: list[dict] = []
        for row in rows[1:]:
            if not any(c is not None and str(c).strip() != "" for c in row):
                continue
            d: dict[str, Any] = {}
            for idx, val in enumerate(row):
                k = col_map.get(idx)
                if not k:
                    continue
                d[k] = val
            if d:
                out.append(d)
        return out

    @staticmethod
    async def import_excel(
        db: AsyncSession,
        *,
        file_name: str,
        file_bytes: bytes,
        current_user_id: int,
    ) -> WaybillImportBatch:
        """主入口：解析 + 落库 + 同步逐行 create_waybill。"""
        try:
            rows = WaybillImportService._parse_excel(file_bytes)
        except BizException:
            raise
        except Exception as e:  # pragma: no cover
            raise BizException(f"Excel 解析失败：{e}") from e

        batch = WaybillImportBatch(
            file_name=file_name,
            total_count=len(rows),
            success_count=0,
            fail_count=0,
            calc_success_count=0,
            calc_exception_count=0,
            status="importing" if rows else "done",
            created_by=current_user_id,
        )
        db.add(batch)
        await db.flush()

        success = 0
        failed = 0
        for idx, row in enumerate(rows, start=2):  # 表头 1，数据从 2 开始
            row_record = WaybillImportRow(
                batch_id=batch.id,
                row_no=idx,
                raw_data_json=_jsonable(row),
                validate_status="pending",
            )
            db.add(row_record)
            await db.flush()

            try:
                payload = await WaybillImportService._build_waybill_create_from_row(
                    db, row, row_no=idx,
                )
                waybill, _ = await WaybillService.create_waybill(
                    db, payload, current_user_id
                )
                row_record.waybill_id = waybill.id
                row_record.validate_status = "success"
                row_record.calc_status = waybill.calc_status
                success += 1
            except BizException as e:
                row_record.validate_status = "failed"
                row_record.validate_message = str(e)[:1000]
                failed += 1
            except Exception as e:  # noqa
                row_record.validate_status = "failed"
                row_record.validate_message = f"系统异常：{e}"[:1000]
                failed += 1
            await db.flush()

        batch.success_count = success
        batch.fail_count = failed
        batch.status = "done" if failed == 0 else (
            "imported" if success > 0 else "failed"
        )
        await db.flush()
        await db.commit()
        return batch

    @staticmethod
    async def get_batch(
        db: AsyncSession, batch_id: int
    ) -> Optional[WaybillImportBatch]:
        r = await db.execute(
            select(WaybillImportBatch).where(
                WaybillImportBatch.id == batch_id,
                WaybillImportBatch.is_deleted == 0,
            )
        )
        return r.scalar_one_or_none()

    @staticmethod
    async def page_batches(
        db: AsyncSession, *, page: int = 1, limit: int = 20
    ) -> dict:
        from sqlalchemy import func as _f
        offset = (page - 1) * limit
        cnt = await db.execute(
            select(_f.count(WaybillImportBatch.id)).where(
                WaybillImportBatch.is_deleted == 0
            )
        )
        total = int(cnt.scalar_one() or 0)
        r = await db.execute(
            select(WaybillImportBatch).where(
                WaybillImportBatch.is_deleted == 0,
            ).order_by(WaybillImportBatch.id.desc())
            .offset(offset).limit(limit)
        )
        items = []
        for b in r.scalars().all():
            items.append({
                "id": b.id,
                "fileName": b.file_name,
                "totalCount": b.total_count,
                "successCount": b.success_count,
                "failCount": b.fail_count,
                "calcSuccessCount": b.calc_success_count,
                "calcExceptionCount": b.calc_exception_count,
                "status": b.status,
                "errorMessage": b.error_message,
                "createdBy": b.created_by,
                "createdAt": b.created_at,
            })
        return {"list": items, "total": total, "page": page, "limit": limit}

    @staticmethod
    async def list_rows(
        db: AsyncSession, batch_id: int,
        *, validate_status: Optional[str] = None,
        page: int = 1, limit: int = 50,
    ) -> dict:
        from sqlalchemy import func as _f
        q = select(WaybillImportRow).where(
            WaybillImportRow.batch_id == batch_id,
            WaybillImportRow.is_deleted == 0,
        )
        cq = select(_f.count(WaybillImportRow.id)).where(
            WaybillImportRow.batch_id == batch_id,
            WaybillImportRow.is_deleted == 0,
        )
        if validate_status:
            q = q.where(WaybillImportRow.validate_status == validate_status)
            cq = cq.where(WaybillImportRow.validate_status == validate_status)
        cnt = await db.execute(cq)
        total = int(cnt.scalar_one() or 0)
        offset = (page - 1) * limit
        r = await db.execute(
            q.order_by(WaybillImportRow.row_no.asc())
            .offset(offset).limit(limit)
        )
        items = []
        for row in r.scalars().all():
            items.append({
                "id": row.id,
                "batchId": row.batch_id,
                "rowNo": row.row_no,
                "rawData": row.raw_data_json,
                "validateStatus": row.validate_status,
                "validateMessage": row.validate_message,
                "waybillId": row.waybill_id,
                "calcStatus": row.calc_status,
                "createdAt": row.created_at,
            })
        return {"list": items, "total": total, "page": page, "limit": limit}


def _jsonable(d: dict) -> dict:
    """将解析行中可能的 datetime/Decimal 转为 JSON-safe。"""
    out: dict[str, Any] = {}
    for k, v in d.items():
        if v is None:
            out[k] = None
        elif hasattr(v, "isoformat"):
            out[k] = v.isoformat()
        elif isinstance(v, (str, int, float, bool)):
            out[k] = v
        else:
            out[k] = str(v)
    return out
