"""经营核算 Service（文档 13）

**财务确认口径**的收入成本毛利，与 ``insight.ProfitService`` 的理论值口径并存：
驾驶舱看趋势（第二天就有数，月底会变），核算对账本（滞后于对账进度，出来就不再变）。
两者数字不同是设计使然，页面上必须写清差异，不要试图对平。

实现要点：

- **不建汇总表**，全部按期间实时聚合（文档 13 §5.1）。单租户月度单据量在千级。
- **归期用财务期间**：对账单/结算单/工资单按 ``period_end`` 月份，任务费用单按
  ``actual_pay_time`` 月份。不用业务发生时间，这是与驾驶舱数字不同的主因。
- **预付单不计成本**：预付是资金流出不是成本发生，承运商结算单金额已扣预付净额，
  两处都计会翻倍。
- **成本先摊到任务、再摊到运单**：结算单/工资单覆盖多个任务，按任务侧净额拆到任务；
  任务再按挂接台数拆到运单。摊不掉的（空驶、无挂接行）显式进「未分摊成本」。
"""

from calendar import monthrange
from datetime import date as ddate, datetime
from decimal import ROUND_HALF_UP, Decimal
from typing import Any, Dict, List, Optional, Sequence, Tuple

from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.common.exceptions import BizException
from app.modules.client.models.finance.carrier_recon import CarrierReconTaskLink
from app.modules.client.models.finance.carrier_settlement_doc import (
    CarrierSettlementDoc,
    CarrierSettleReconLink,
)
from app.modules.client.models.finance.customer_invoice import CustomerInvoice
from app.modules.client.models.finance.customer_recon import (
    CustomerRecon,
    CustomerReconWaybillLink,
)
from app.modules.client.models.finance.customer_settlement import CustomerSettlement
from app.modules.client.models.finance.driver_payroll import (
    DriverPayroll,
    DriverPayrollTaskLink,
)
from app.modules.client.models.finance.vendor_invoice import VendorInvoiceSettleLink
from app.modules.client.models.task.task import Task
from app.modules.client.models.task.task_finance_doc import TaskFinanceDoc
from app.modules.client.models.task.task_waybill_item import TaskWaybillItem
from app.modules.client.models.waybill.waybill import Waybill
from app.modules.client.services.finance.accounting.accounting_constants import (
    ALLOCATED_DIMENSIONS,
    CARRIER_TYPE_LABELS,
    COST_CARRIER_SETTLE_STATUSES,
    COST_PAYROLL_STATUSES,
    COST_TASK_FINANCE_DOC_TYPES,
    COST_TASK_FINANCE_STATUSES,
    DEFAULT_INPUT_TAX_RATE,
    DEFAULT_OUTPUT_TAX_RATE,
    DIM_CARRIER_TYPE,
    DIM_CUSTOMER,
    DIM_DRIVER,
    DIM_ENTITY,
    DIM_ROUTE,
    DIM_VEHICLE,
    DIMENSION_LABELS,
    DIMENSIONS,
    OUTPUT_TAX_RATE_CONFIG_GROUP,
    OUTPUT_TAX_RATE_CONFIG_KEY,
    OUTPUT_TAX_RATE_DESCRIPTION,
    REVENUE_RECON_STATUSES,
    REVENUE_SETTLE_STATUSES,
    TAX_MODE_EXCL,
    TAX_MODES,
    UNALLOCATED_KEY,
    UNALLOCATED_LABEL,
)
from app.modules.client.services.finance.accounting.cost_allocator import (
    allocate_task_cost_to_waybills,
    split_doc_amount_by_task,
)
from app.modules.client.services.system_config_service import SystemConfigService

_CENT = Decimal("0.01")
_ZERO = Decimal("0")


class ProfitAccountingService:
    """财务确认口径的经营核算"""

    # ------------------------------------------------------------------
    # 期间
    # ------------------------------------------------------------------
    @staticmethod
    def parse_period(period: Optional[str]) -> Tuple[ddate, ddate, str]:
        """把 ``YYYY-MM`` / ``YYYY-Qn`` / ``YYYY`` 解析成起止日期。

        核算是按期看的，所以不接受任意起止日期——「6 月 3 日到 7 月 12 日的毛利」在
        财务口径下没有意义。
        """
        text = (period or "").strip()
        if not text:
            today = ddate.today()
            text = today.strftime("%Y-%m")
        parts = text.upper().split("-")
        try:
            year = int(parts[0])
        except (ValueError, IndexError):
            raise BizException("期间格式不对，请填 2026-06、2026-Q2 或 2026")
        if len(parts) == 1:
            return ddate(year, 1, 1), ddate(year, 12, 31), f"{year} 年"
        token = parts[1]
        if token.startswith("Q"):
            try:
                quarter = int(token[1:])
            except ValueError:
                raise BizException("季度格式不对，请填 2026-Q2 这样的写法")
            if not 1 <= quarter <= 4:
                raise BizException("季度只能是 Q1 到 Q4")
            start_month = (quarter - 1) * 3 + 1
            end_month = start_month + 2
            return (
                ddate(year, start_month, 1),
                ddate(year, end_month, monthrange(year, end_month)[1]),
                f"{year} 年第 {quarter} 季度",
            )
        try:
            month = int(token)
        except ValueError:
            raise BizException("期间格式不对，请填 2026-06、2026-Q2 或 2026")
        if not 1 <= month <= 12:
            raise BizException("月份只能是 1 到 12")
        return (
            ddate(year, month, 1),
            ddate(year, month, monthrange(year, month)[1]),
            f"{year} 年 {month} 月",
        )

    # ------------------------------------------------------------------
    # KPI
    # ------------------------------------------------------------------
    @classmethod
    async def kpi(
        cls,
        db: AsyncSession,
        *,
        period: Optional[str] = None,
        enterprise_id: Optional[int] = None,
        tax_mode: str = TAX_MODE_EXCL,
    ) -> dict:
        """期间 KPI：收入、成本、毛利、毛利率、缺票税损、未分摊成本。"""
        cls._assert_tax_mode(tax_mode)
        start, end, label = cls.parse_period(period)
        revenue = await cls._revenue_lines(
            db, start, end, enterprise_id=enterprise_id,
        )
        realized = await cls._realized_revenue(
            db, start, end, enterprise_id=enterprise_id,
        )
        costs = await cls._cost_docs(db, start, end, enterprise_id=enterprise_id)
        rate = await cls.default_output_tax_rate(db)
        rate_by_customer = await cls._output_rate_by_customer(db, start, end)

        confirmed = _money(sum((r["amount"] for r in revenue), _ZERO))
        revenue_excl = _money(sum(
            (
                _excl(r["amount"], rate_by_customer.get(r["customerId"], rate))
                for r in revenue
            ),
            _ZERO,
        ))
        cost_incl = _money(sum((c["amount"] for c in costs), _ZERO))
        cost_excl = _money(sum((c["amountExclTax"] for c in costs), _ZERO))
        no_invoice = _money(sum((c["uninvoicedAmount"] for c in costs), _ZERO))
        tax_loss = _money(
            no_invoice * Decimal(DEFAULT_INPUT_TAX_RATE)
            / (Decimal(100) + Decimal(DEFAULT_INPUT_TAX_RATE))
        )
        _, unallocated = await cls._allocate_costs(db, costs)

        gross_incl = _money(confirmed - cost_incl)
        gross_excl = _money(revenue_excl - cost_excl)
        margin_base = revenue_excl if tax_mode == TAX_MODE_EXCL else confirmed
        margin_value = gross_excl if tax_mode == TAX_MODE_EXCL else gross_incl
        return {
            "period": period or start.strftime("%Y-%m"),
            "periodLabel": label,
            "periodStart": start,
            "periodEnd": end,
            "taxMode": tax_mode,
            "outputTaxRate": float(rate),
            "confirmedRevenue": float(confirmed),
            "realizedRevenue": float(realized),
            "revenueExclTax": float(revenue_excl),
            "costInclTax": float(cost_incl),
            "costExclTax": float(cost_excl),
            "grossProfitInclTax": float(gross_incl),
            "grossProfitExclTax": float(gross_excl),
            "grossMarginRate": (
                round(float(margin_value / margin_base * 100), 2)
                if margin_base > 0 else None
            ),
            "noInvoiceCost": float(no_invoice),
            "missingInvoiceTaxLoss": float(tax_loss),
            "unallocatedCost": float(unallocated),
            "revenueDocCount": len({r["reconId"] for r in revenue}),
            "costDocCount": len(costs),
        }

    # ------------------------------------------------------------------
    # 维度汇总
    # ------------------------------------------------------------------
    @classmethod
    async def by_dimension(
        cls,
        db: AsyncSession,
        *,
        dimension: str,
        period: Optional[str] = None,
        enterprise_id: Optional[int] = None,
        tax_mode: str = TAX_MODE_EXCL,
    ) -> List[dict]:
        """按维度汇总收入、成本、毛利、毛利率。

        「未分摊成本」始终作为独立一行返回（收入为 0），不并进任何业务维度值。
        """
        cls._assert_tax_mode(tax_mode)
        if dimension not in DIMENSIONS:
            raise BizException(
                "维度不支持，请选择客户、经营主体、线路、车辆、司机或承运类型"
            )
        start, end, _ = cls.parse_period(period)
        revenue = await cls._revenue_lines(
            db, start, end, enterprise_id=enterprise_id,
        )
        costs = await cls._cost_docs(db, start, end, enterprise_id=enterprise_id)
        rate = await cls.default_output_tax_rate(db)
        rate_by_customer = await cls._output_rate_by_customer(db, start, end)

        buckets: Dict[str, dict] = {}

        def slot(key: Any, name: Optional[str]) -> dict:
            k = str(key)
            row = buckets.setdefault(k, {
                "dimension": dimension,
                "dimensionValue": k,
                "dimensionLabel": name or k,
                "revenue": _ZERO,
                "revenueExclTax": _ZERO,
                "cost": _ZERO,
                "costExclTax": _ZERO,
            })
            if name and row["dimensionLabel"] in (None, k):
                row["dimensionLabel"] = name
            return row

        if dimension in ALLOCATED_DIMENSIONS:
            waybill_costs, unallocated = await cls._allocate_costs(db, costs)
            excl_ratio = cls._excl_ratio(costs)
            meta = await cls._waybill_meta(db, list(waybill_costs.keys()))
            for line in revenue:
                key, name = cls._revenue_key(dimension, line, meta)
                row = slot(key, name)
                row["revenue"] += line["amount"]
                row["revenueExclTax"] += _excl(
                    line["amount"],
                    rate_by_customer.get(line["customerId"], rate),
                )
            for waybill_id, amount in waybill_costs.items():
                info = meta.get(int(waybill_id)) or {}
                key, name = cls._waybill_key(dimension, info)
                row = slot(key, name)
                row["cost"] += amount
                row["costExclTax"] += _money(amount * excl_ratio)
            if unallocated > 0:
                row = slot(UNALLOCATED_KEY, UNALLOCATED_LABEL)
                row["cost"] += unallocated
                row["costExclTax"] += _money(unallocated * excl_ratio)
        else:
            if dimension == DIM_ENTITY:
                for line in revenue:
                    row = slot(line["enterpriseId"] or 0, None)
                    row["revenue"] += line["amount"]
                    row["revenueExclTax"] += _excl(
                        line["amount"],
                        rate_by_customer.get(line["customerId"], rate),
                    )
            for cost in costs:
                key, name = await cls._cost_key(db, dimension, cost)
                row = slot(key, name)
                row["cost"] += cost["amount"]
                row["costExclTax"] += cost["amountExclTax"]

        rows: List[dict] = []
        for row in buckets.values():
            revenue_amt = _money(row["revenue"])
            revenue_excl = _money(row["revenueExclTax"])
            cost_amt = _money(row["cost"])
            cost_excl = _money(row["costExclTax"])
            gross = (
                _money(revenue_excl - cost_excl) if tax_mode == TAX_MODE_EXCL
                else _money(revenue_amt - cost_amt)
            )
            base = revenue_excl if tax_mode == TAX_MODE_EXCL else revenue_amt
            rows.append({
                "dimension": row["dimension"],
                "dimensionValue": row["dimensionValue"],
                "dimensionLabel": row["dimensionLabel"],
                "revenue": float(revenue_amt),
                "revenueExclTax": float(revenue_excl),
                "cost": float(cost_amt),
                "costExclTax": float(cost_excl),
                "grossProfit": float(gross),
                "grossMarginRate": (
                    round(float(gross / base * 100), 2) if base > 0 else None
                ),
            })
        rows.sort(key=lambda x: x["revenue"] or -x["cost"], reverse=True)
        return rows

    # ------------------------------------------------------------------
    # 下钻
    # ------------------------------------------------------------------
    @classmethod
    async def drill_down(
        cls,
        db: AsyncSession,
        *,
        dimension: str,
        dimension_value: str,
        period: Optional[str] = None,
        enterprise_id: Optional[int] = None,
    ) -> dict:
        """某维度值下的收入单据与成本单据清单。"""
        if dimension not in DIMENSIONS:
            raise BizException("维度不支持，请重新选择")
        start, end, label = cls.parse_period(period)
        revenue = await cls._revenue_lines(
            db, start, end, enterprise_id=enterprise_id,
        )
        costs = await cls._cost_docs(db, start, end, enterprise_id=enterprise_id)

        revenue_rows: List[dict] = []
        cost_rows: List[dict] = []
        target = str(dimension_value)

        if dimension in ALLOCATED_DIMENSIONS:
            waybill_costs, unallocated = await cls._allocate_costs(db, costs)
            meta = await cls._waybill_meta(
                db,
                list({*waybill_costs.keys(), *[r["waybillId"] for r in revenue]}),
            )
            for line in revenue:
                key, _ = cls._revenue_key(dimension, line, meta)
                if str(key) == target:
                    revenue_rows.append(cls._revenue_row(line))
            # 成本按单据呈现：先找出落在该维度值下的运单，再回溯它们所属的成本单
            hit_waybills = {
                int(wid) for wid in waybill_costs
                if str(cls._waybill_key(dimension, meta.get(int(wid)) or {})[0])
                == target
            }
            doc_amount = await cls._cost_docs_of_waybills(
                db, costs, hit_waybills,
            ) if hit_waybills else {}
            for cost in costs:
                amount = doc_amount.get((cost["docKind"], cost["docId"]))
                if amount:
                    cost_rows.append(cls._cost_row(cost, allocated=amount))
            if target == UNALLOCATED_KEY and unallocated > 0:
                for cost in costs:
                    if cost.get("unallocatedAmount"):
                        cost_rows.append(
                            cls._cost_row(cost, allocated=cost["unallocatedAmount"])
                        )
        else:
            if dimension == DIM_ENTITY:
                for line in revenue:
                    if str(line["enterpriseId"] or 0) == target:
                        revenue_rows.append(cls._revenue_row(line))
            for cost in costs:
                key, _ = await cls._cost_key(db, dimension, cost)
                if str(key) == target:
                    cost_rows.append(cls._cost_row(cost))

        return {
            "dimension": dimension,
            "dimensionValue": target,
            "dimensionLabel": DIMENSION_LABELS.get(dimension, dimension),
            "periodLabel": label,
            "revenueDocs": revenue_rows,
            "costDocs": cost_rows,
            "revenueTotal": float(_money(sum(
                (Decimal(str(r["amount"])) for r in revenue_rows), _ZERO,
            ))),
            "costTotal": float(_money(sum(
                (Decimal(str(r["amount"])) for r in cost_rows), _ZERO,
            ))),
        }

    # ------------------------------------------------------------------
    # 跨主体待结转（文档 13 §六，只呈现不结算）
    # ------------------------------------------------------------------
    @classmethod
    async def inter_entity(
        cls,
        db: AsyncSession,
        *,
        period: Optional[str] = None,
    ) -> dict:
        """跨主体明细：运单归属主体 ≠ 执行任务归属主体的部分。

        本期只呈现规模，不生成主体间结算单据（定价口径未定，见文档 13 §6.3）。
        """
        start, end, label = cls.parse_period(period)
        revenue = await cls._revenue_lines(db, start, end)
        waybill_ids = [r["waybillId"] for r in revenue if r["waybillId"]]
        if not waybill_ids:
            return {
                "periodLabel": label, "rows": [], "byEntity": [],
            }
        r = await db.execute(
            select(
                TaskWaybillItem.waybill_id,
                TaskWaybillItem.task_id,
                TaskWaybillItem.quantity,
                Task.enterprise_id,
                Task.task_no,
                Waybill.enterprise_id,
                Waybill.waybill_no,
            )
            .join(Task, Task.id == TaskWaybillItem.task_id)
            .join(Waybill, Waybill.id == TaskWaybillItem.waybill_id)
            .where(
                TaskWaybillItem.waybill_id.in_(list(set(waybill_ids))),
                TaskWaybillItem.is_deleted == 0,
                Task.is_deleted == 0,
                Waybill.is_deleted == 0,
            )
        )
        revenue_by_waybill: Dict[int, Decimal] = {}
        for line in revenue:
            wid = int(line["waybillId"] or 0)
            revenue_by_waybill[wid] = (
                revenue_by_waybill.get(wid, _ZERO) + line["amount"]
            )

        rows: List[dict] = []
        by_entity: Dict[int, dict] = {}
        for wid, task_id, qty, task_entity, task_no, wb_entity, wb_no in r.all():
            if not task_entity or not wb_entity or int(task_entity) == int(wb_entity):
                continue
            amount = revenue_by_waybill.get(int(wid), _ZERO)
            rows.append({
                "waybillId": int(wid),
                "waybillNo": wb_no,
                "taskId": int(task_id),
                "taskNo": task_no,
                "quantity": float(qty or 0),
                "revenueEntityId": int(wb_entity),
                "costEntityId": int(task_entity),
                "revenueAmount": float(_money(amount)),
            })
            out = by_entity.setdefault(int(task_entity), {
                "enterpriseId": int(task_entity),
                "transferOutAmount": 0.0,
                "transferInAmount": 0.0,
                "count": 0,
            })
            out["transferOutAmount"] += float(_money(amount))
            out["count"] += 1
            inbound = by_entity.setdefault(int(wb_entity), {
                "enterpriseId": int(wb_entity),
                "transferOutAmount": 0.0,
                "transferInAmount": 0.0,
                "count": 0,
            })
            inbound["transferInAmount"] += float(_money(amount))
        return {
            "periodLabel": label,
            "rows": rows,
            "byEntity": list(by_entity.values()),
        }

    # ------------------------------------------------------------------
    # 底稿导出
    # ------------------------------------------------------------------
    @classmethod
    async def build_export_workbook(
        cls,
        db: AsyncSession,
        *,
        period: Optional[str] = None,
        enterprise_id: Optional[int] = None,
    ) -> bytes:
        """底稿：收入单据与成本单据两张表页，供财务逐笔核对。"""
        try:
            import openpyxl  # type: ignore
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError as e:  # pragma: no cover
            raise BizException("导出失败，服务端缺少表格组件，请联系管理员") from e

        start, end, label = cls.parse_period(period)
        revenue = await cls._revenue_lines(
            db, start, end, enterprise_id=enterprise_id,
        )
        costs = await cls._cost_docs(db, start, end, enterprise_id=enterprise_id)
        caption = (
            f"核算期间 {label}（{start.isoformat()} ~ {end.isoformat()}）；"
            "口径：已确认对账收入 / 已审批成本，预付单不计成本"
        )
        head_fill = PatternFill("solid", fgColor="FFE7EEF8")
        head_font = Font(bold=True, size=11)

        def _head(ws, headers: Sequence[str], widths: Sequence[int]) -> None:
            ws.append([caption])
            ws.append(list(headers))
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=2, column=col)
                cell.fill = head_fill
                cell.font = head_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = w

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "收入底稿"
        _head(
            ws,
            ["对账单号", "客户", "经营主体", "运单号", "行金额", "调整额", "对账周期止"],
            [18, 24, 12, 18, 14, 12, 14],
        )
        for line in revenue:
            ws.append([
                line["docNo"], line["customerName"], line["enterpriseId"],
                line["waybillNo"], float(line["amount"]),
                float(line["adjustAmount"]), line["periodEnd"],
            ])

        ws2 = wb.create_sheet("成本底稿")
        _head(
            ws2,
            [
                "单据类型", "单号", "对象", "经营主体", "含税金额",
                "不含税金额", "无票金额", "归期",
            ],
            [16, 18, 24, 12, 14, 14, 14, 14],
        )
        for cost in costs:
            ws2.append([
                cost["docKindLabel"], cost["docNo"], cost["payeeName"],
                cost["enterpriseId"], float(cost["amount"]),
                float(cost["amountExclTax"]), float(cost["uninvoicedAmount"]),
                cost["periodEnd"],
            ])

        from io import BytesIO

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    # ------------------------------------------------------------------
    # 税率
    # ------------------------------------------------------------------
    @classmethod
    async def default_output_tax_rate(cls, db: AsyncSession) -> Decimal:
        """默认销项税率（百分数）。缺配置时懒补齐，避免存量租户读到空。"""
        await SystemConfigService.ensure(
            db,
            OUTPUT_TAX_RATE_CONFIG_KEY,
            default_value=str(DEFAULT_OUTPUT_TAX_RATE),
            config_group=OUTPUT_TAX_RATE_CONFIG_GROUP,
            description=OUTPUT_TAX_RATE_DESCRIPTION,
            value_type="number",
        )
        raw = await SystemConfigService.get_by_key(db, OUTPUT_TAX_RATE_CONFIG_KEY)
        try:
            rate = Decimal(str(raw))
        except Exception:
            rate = Decimal(DEFAULT_OUTPUT_TAX_RATE)
        if rate < 0 or rate > 100:
            rate = Decimal(DEFAULT_OUTPUT_TAX_RATE)
        return rate

    # ------------------------------------------------------------------
    # 取数：收入
    # ------------------------------------------------------------------
    @classmethod
    async def _revenue_lines(
        cls,
        db: AsyncSession,
        start: ddate,
        end: ddate,
        *,
        enterprise_id: Optional[int] = None,
    ) -> List[dict]:
        """已确认收入的对账行明细（归期按对账单 ``period_end``）。"""
        stmt = (
            select(
                CustomerRecon.id,
                CustomerRecon.doc_no,
                CustomerRecon.customer_id,
                CustomerRecon.customer_name,
                CustomerRecon.enterprise_id,
                CustomerRecon.period_end,
                CustomerReconWaybillLink.waybill_id,
                CustomerReconWaybillLink.waybill_no,
                CustomerReconWaybillLink.amount,
                CustomerReconWaybillLink.adjust_amount,
            )
            .join(
                CustomerReconWaybillLink,
                CustomerReconWaybillLink.recon_id == CustomerRecon.id,
            )
            .where(
                CustomerRecon.is_deleted == 0,
                CustomerRecon.status.in_(REVENUE_RECON_STATUSES),
                CustomerReconWaybillLink.is_deleted == 0,
                CustomerRecon.period_end >= _day_start(start),
                CustomerRecon.period_end <= _day_end(end),
            )
        )
        if enterprise_id:
            stmt = stmt.where(CustomerRecon.enterprise_id == enterprise_id)
        r = await db.execute(stmt)
        return [
            {
                "reconId": int(row[0]),
                "docNo": row[1],
                "customerId": row[2],
                "customerName": row[3],
                "enterpriseId": row[4],
                "periodEnd": row[5],
                "waybillId": row[6],
                "waybillNo": row[7],
                "amount": Decimal(str(row[8] or 0)),
                "adjustAmount": Decimal(str(row[9] or 0)),
            }
            for row in r.all()
        ]

    @classmethod
    async def _realized_revenue(
        cls,
        db: AsyncSession,
        start: ddate,
        end: ddate,
        *,
        enterprise_id: Optional[int] = None,
    ) -> Decimal:
        """已实现收入：期间内已收款结算单的实收合计（收付实现制口径）。"""
        stmt = select(
            func.coalesce(func.sum(CustomerSettlement.actual_amount), 0)
        ).where(
            CustomerSettlement.is_deleted == 0,
            CustomerSettlement.status.in_(REVENUE_SETTLE_STATUSES),
            CustomerSettlement.received_at >= _day_start(start),
            CustomerSettlement.received_at <= _day_end(end),
        )
        if enterprise_id:
            stmt = stmt.where(CustomerSettlement.enterprise_id == enterprise_id)
        return _money(Decimal(str((await db.execute(stmt)).scalar() or 0)))

    @classmethod
    async def _output_rate_by_customer(
        cls, db: AsyncSession, start: ddate, end: ddate,
    ) -> Dict[int, Decimal]:
        """按客户取期间内已开票的加权平均税率（无票客户不在返回里）。"""
        r = await db.execute(
            select(
                CustomerInvoice.customer_id,
                func.sum(CustomerInvoice.amount_incl_tax),
                func.sum(CustomerInvoice.tax_amount),
            ).where(
                CustomerInvoice.is_deleted == 0,
                CustomerInvoice.status == 3,
                CustomerInvoice.invoice_date >= start,
                CustomerInvoice.invoice_date <= end,
            ).group_by(CustomerInvoice.customer_id)
        )
        out: Dict[int, Decimal] = {}
        for customer_id, incl, tax in r.all():
            incl_d = Decimal(str(incl or 0))
            tax_d = Decimal(str(tax or 0))
            excl = incl_d - tax_d
            if excl > 0 and tax_d > 0:
                out[int(customer_id)] = _money(tax_d / excl * 100)
        return out

    # ------------------------------------------------------------------
    # 取数：成本
    # ------------------------------------------------------------------
    @classmethod
    async def _cost_docs(
        cls,
        db: AsyncSession,
        start: ddate,
        end: ddate,
        *,
        enterprise_id: Optional[int] = None,
    ) -> List[dict]:
        """三条成本线的单据清单（含不含税金额与无票金额）。"""
        out: List[dict] = []

        # --- 承运商结算单：归期按 period_end，不含税按已核销进项票 ---
        stmt = select(CarrierSettlementDoc).where(
            CarrierSettlementDoc.is_deleted == 0,
            CarrierSettlementDoc.status.in_(COST_CARRIER_SETTLE_STATUSES),
            CarrierSettlementDoc.period_end >= _day_start(start),
            CarrierSettlementDoc.period_end <= _day_end(end),
        )
        if enterprise_id:
            stmt = stmt.where(CarrierSettlementDoc.enterprise_id == enterprise_id)
        settles = list((await db.execute(stmt)).scalars().all())
        excl_map = await cls._settle_invoice_excl(
            db, [int(x.id) for x in settles],
        )
        for settle in settles:
            amount = Decimal(str(settle.planned_amount or 0))
            invoiced_incl, invoiced_excl = excl_map.get(
                int(settle.id), (_ZERO, _ZERO),
            )
            uninvoiced = max(amount - invoiced_incl, _ZERO)
            out.append({
                "docKind": "carrier_settle",
                "docKindLabel": "承运商结算单",
                "docId": int(settle.id),
                "docNo": settle.doc_no,
                "payeeId": settle.carrier_id,
                "payeeName": settle.carrier_name,
                "enterpriseId": settle.enterprise_id,
                "periodEnd": settle.period_end,
                "amount": _money(amount),
                # 有票部分按票面不含税，无票部分全额是成本（不能抵扣）
                "amountExclTax": _money(invoiced_excl + uninvoiced),
                "uninvoicedAmount": _money(uninvoiced),
            })

        # --- 司机工资单：人力成本没有进项票，含税=不含税 ---
        stmt = select(DriverPayroll).where(
            DriverPayroll.is_deleted == 0,
            DriverPayroll.status.in_(COST_PAYROLL_STATUSES),
            DriverPayroll.period_end >= _day_start(start),
            DriverPayroll.period_end <= _day_end(end),
        )
        if enterprise_id:
            stmt = stmt.where(DriverPayroll.enterprise_id == enterprise_id)
        for payroll in (await db.execute(stmt)).scalars().all():
            amount = Decimal(str(payroll.net_amount or 0))
            out.append({
                "docKind": "driver_payroll",
                "docKindLabel": "司机工资单",
                "docId": int(payroll.id),
                "docNo": payroll.doc_no,
                "payeeId": payroll.driver_id,
                "payeeName": payroll.driver_name,
                "enterpriseId": payroll.enterprise_id,
                "periodEnd": payroll.period_end,
                "amount": _money(amount),
                "amountExclTax": _money(amount),
                "uninvoicedAmount": _ZERO,
            })

        # --- 任务级费用单：归期按 actual_pay_time，预付单不计 ---
        stmt = select(TaskFinanceDoc).where(
            TaskFinanceDoc.is_deleted == 0,
            TaskFinanceDoc.status.in_(COST_TASK_FINANCE_STATUSES),
            TaskFinanceDoc.doc_type.in_(COST_TASK_FINANCE_DOC_TYPES),
            TaskFinanceDoc.actual_pay_time >= _day_start(start),
            TaskFinanceDoc.actual_pay_time <= _day_end(end),
        )
        if enterprise_id:
            stmt = stmt.where(TaskFinanceDoc.enterprise_id == enterprise_id)
        for doc in (await db.execute(stmt)).scalars().all():
            amount = Decimal(str(doc.actual_amount or doc.planned_amount or 0))
            out.append({
                "docKind": "task_finance",
                "docKindLabel": "任务费用单",
                "docId": int(doc.id),
                "docNo": doc.doc_no,
                "payeeId": doc.payee_id,
                "payeeName": doc.payee_name,
                "payeeType": int(doc.payee_type or 0),
                "enterpriseId": doc.enterprise_id,
                "periodEnd": doc.actual_pay_time,
                "taskId": doc.task_id,
                "amount": _money(amount),
                "amountExclTax": _money(amount),
                "uninvoicedAmount": _money(amount),
            })
        return out

    @classmethod
    async def _settle_invoice_excl(
        cls, db: AsyncSession, settle_ids: Sequence[int],
    ) -> Dict[int, Tuple[Decimal, Decimal]]:
        """每张结算单已核销进项票的 ``(含税额, 不含税额)``。"""
        if not settle_ids:
            return {}
        from app.modules.client.models.finance.vendor_invoice import VendorInvoice

        r = await db.execute(
            select(
                VendorInvoiceSettleLink.settle_id,
                VendorInvoiceSettleLink.applied_amount,
                VendorInvoice.amount_incl_tax,
                VendorInvoice.amount_excl_tax,
            )
            .join(
                VendorInvoice,
                VendorInvoice.id == VendorInvoiceSettleLink.invoice_id,
            )
            .where(
                VendorInvoiceSettleLink.settle_id.in_(list(settle_ids)),
                VendorInvoiceSettleLink.is_deleted == 0,
                VendorInvoice.is_deleted == 0,
                VendorInvoice.status.notin_((4, 9)),
            )
        )
        out: Dict[int, Tuple[Decimal, Decimal]] = {}
        for settle_id, applied, incl, excl in r.all():
            applied_d = Decimal(str(applied or 0))
            incl_d = Decimal(str(incl or 0))
            excl_d = Decimal(str(excl or 0))
            # 核销额可能只是票面的一部分，按比例折算不含税额
            share = (excl_d / incl_d) if incl_d > 0 else Decimal("1")
            cur = out.get(int(settle_id), (_ZERO, _ZERO))
            out[int(settle_id)] = (
                cur[0] + applied_d,
                cur[1] + _money(applied_d * share),
            )
        return out

    # ------------------------------------------------------------------
    # 分摊
    # ------------------------------------------------------------------
    @classmethod
    async def _allocate_costs(
        cls, db: AsyncSession, costs: Sequence[dict],
    ) -> Tuple[Dict[int, Decimal], Decimal]:
        """成本单 → 任务 → 运单两级分摊，返回 ``({运单: 成本}, 未分摊合计)``。"""
        task_costs: Dict[int, Decimal] = {}
        unallocated = _ZERO
        for cost in costs:
            weights = await cls._task_weights(db, cost)
            if not weights:
                cost["unallocatedAmount"] = cost["amount"]
                unallocated += cost["amount"]
                continue
            for task_id, amount in split_doc_amount_by_task(
                cost["amount"], weights,
            ):
                task_costs[task_id] = task_costs.get(task_id, _ZERO) + amount

        items = await cls._task_items(db, list(task_costs.keys()))
        allocated, task_unallocated = allocate_task_cost_to_waybills(
            task_costs, items,
        )
        return allocated, _money(unallocated + task_unallocated)

    @classmethod
    async def _task_weights(
        cls, db: AsyncSession, cost: dict,
    ) -> List[Tuple[int, Decimal]]:
        """一张成本单覆盖的任务及其权重（任务侧净额）。"""
        kind = cost["docKind"]
        if kind == "task_finance":
            task_id = cost.get("taskId")
            return [(int(task_id), cost["amount"])] if task_id else []
        if kind == "carrier_settle":
            r = await db.execute(
                select(
                    CarrierReconTaskLink.task_id,
                    func.coalesce(func.sum(CarrierReconTaskLink.net_amount), 0),
                )
                .join(
                    CarrierSettleReconLink,
                    CarrierSettleReconLink.recon_id == CarrierReconTaskLink.recon_id,
                )
                .where(
                    CarrierSettleReconLink.settle_id == cost["docId"],
                    CarrierSettleReconLink.is_deleted == 0,
                    CarrierReconTaskLink.is_deleted == 0,
                )
                .group_by(CarrierReconTaskLink.task_id)
            )
            return [(int(t), Decimal(str(w or 0))) for t, w in r.all()]
        r = await db.execute(
            select(
                DriverPayrollTaskLink.task_id,
                func.coalesce(func.sum(DriverPayrollTaskLink.commission_amount), 0),
            ).where(
                DriverPayrollTaskLink.payroll_id == cost["docId"],
                DriverPayrollTaskLink.is_deleted == 0,
            ).group_by(DriverPayrollTaskLink.task_id)
        )
        return [(int(t), Decimal(str(w or 0))) for t, w in r.all()]

    @classmethod
    async def _task_items(
        cls, db: AsyncSession, task_ids: Sequence[int],
    ) -> Dict[int, List[Tuple[int, Decimal]]]:
        if not task_ids:
            return {}
        r = await db.execute(
            select(
                TaskWaybillItem.task_id,
                TaskWaybillItem.waybill_id,
                func.coalesce(func.sum(TaskWaybillItem.quantity), 0),
            ).where(
                TaskWaybillItem.task_id.in_(list(task_ids)),
                TaskWaybillItem.is_deleted == 0,
            ).group_by(TaskWaybillItem.task_id, TaskWaybillItem.waybill_id)
        )
        out: Dict[int, List[Tuple[int, Decimal]]] = {}
        for task_id, waybill_id, qty in r.all():
            out.setdefault(int(task_id), []).append(
                (int(waybill_id), Decimal(str(qty or 0)))
            )
        return out

    @classmethod
    async def _cost_docs_of_waybills(
        cls,
        db: AsyncSession,
        costs: Sequence[dict],
        waybill_ids: set,
    ) -> Dict[Tuple[str, int], Decimal]:
        """下钻用：算出每张成本单落在这批运单上的金额。"""
        out: Dict[Tuple[str, int], Decimal] = {}
        for cost in costs:
            weights = await cls._task_weights(db, cost)
            if not weights:
                continue
            per_task = split_doc_amount_by_task(cost["amount"], weights)
            items = await cls._task_items(db, [t for t, _ in per_task])
            allocated, _ = allocate_task_cost_to_waybills(dict(per_task), items)
            hit = sum(
                (amount for wid, amount in allocated.items() if wid in waybill_ids),
                _ZERO,
            )
            if hit > 0:
                out[(cost["docKind"], cost["docId"])] = _money(hit)
        return out

    @staticmethod
    def _excl_ratio(costs: Sequence[dict]) -> Decimal:
        """整体不含税占比：分摊后按这个比例折算各维度的不含税成本。

        逐单分摊两遍（含税一遍、不含税一遍）代价太高，而同期成本的票据结构基本一致，
        用整体比例折算的误差在分位级别，对经营判断没有影响。
        """
        incl = sum((c["amount"] for c in costs), _ZERO)
        excl = sum((c["amountExclTax"] for c in costs), _ZERO)
        return (excl / incl) if incl > 0 else Decimal("1")

    # ------------------------------------------------------------------
    # 维度键
    # ------------------------------------------------------------------
    @classmethod
    async def _waybill_meta(
        cls, db: AsyncSession, waybill_ids: Sequence[int],
    ) -> Dict[int, dict]:
        ids = [int(x) for x in waybill_ids if x]
        if not ids:
            return {}
        r = await db.execute(
            select(
                Waybill.id,
                Waybill.customer_id,
                Waybill.customer_name,
                Waybill.enterprise_id,
                Waybill.origin,
                Waybill.destination,
                Waybill.origin_region_id,
                Waybill.destination_region_id,
            ).where(Waybill.id.in_(ids), Waybill.is_deleted == 0)
        )
        return {
            int(row[0]): {
                "customerId": row[1],
                "customerName": row[2],
                "enterpriseId": row[3],
                "origin": row[4],
                "destination": row[5],
                "originRegionId": row[6],
                "destinationRegionId": row[7],
            }
            for row in r.all()
        }

    @staticmethod
    def _revenue_key(
        dimension: str, line: dict, meta: Dict[int, dict],
    ) -> Tuple[Any, Optional[str]]:
        if dimension == DIM_CUSTOMER:
            return line["customerId"] or 0, line["customerName"]
        info = meta.get(int(line["waybillId"] or 0)) or {}
        return ProfitAccountingService._waybill_key(dimension, info)

    @staticmethod
    def _waybill_key(dimension: str, info: dict) -> Tuple[Any, Optional[str]]:
        if dimension == DIM_CUSTOMER:
            return info.get("customerId") or 0, info.get("customerName")
        if dimension == DIM_ROUTE:
            origin = info.get("origin") or "未知"
            dest = info.get("destination") or "未知"
            return f"{origin}→{dest}", f"{origin} → {dest}"
        return UNALLOCATED_KEY, UNALLOCATED_LABEL

    @classmethod
    async def _cost_key(
        cls, db: AsyncSession, dimension: str, cost: dict,
    ) -> Tuple[Any, Optional[str]]:
        """非分摊维度的成本键：主体直接取单据，车辆/司机/承运类型取任务或收款方。"""
        if dimension == DIM_ENTITY:
            return cost.get("enterpriseId") or 0, None
        if dimension == DIM_DRIVER:
            if cost["docKind"] == "driver_payroll":
                return cost.get("payeeId") or 0, cost.get("payeeName")
            if cost["docKind"] == "task_finance" and int(
                cost.get("payeeType") or 0
            ) == 1:
                return cost.get("payeeId") or 0, cost.get("payeeName")
            return UNALLOCATED_KEY, UNALLOCATED_LABEL
        task_ids = [t for t, _ in await cls._task_weights(db, cost)]
        if not task_ids:
            return UNALLOCATED_KEY, UNALLOCATED_LABEL
        r = await db.execute(
            select(Task.plate_number, Task.carrier_type)
            .where(Task.id == task_ids[0], Task.is_deleted == 0)
        )
        row = r.one_or_none()
        if row is None:
            return UNALLOCATED_KEY, UNALLOCATED_LABEL
        if dimension == DIM_VEHICLE:
            return (row[0] or UNALLOCATED_KEY), (row[0] or UNALLOCATED_LABEL)
        if dimension == DIM_CARRIER_TYPE:
            code = int(row[1] or 0)
            return code, CARRIER_TYPE_LABELS.get(code, "其他")
        return UNALLOCATED_KEY, UNALLOCATED_LABEL

    # ------------------------------------------------------------------
    # 行结构
    # ------------------------------------------------------------------
    @staticmethod
    def _revenue_row(line: dict) -> dict:
        return {
            "docKind": "customer_recon",
            "docKindLabel": "客户对账单",
            "docId": line["reconId"],
            "docNo": line["docNo"],
            "counterparty": line["customerName"],
            "waybillNo": line["waybillNo"],
            "amount": float(line["amount"]),
            "periodEnd": line["periodEnd"],
        }

    @staticmethod
    def _cost_row(cost: dict, *, allocated: Optional[Decimal] = None) -> dict:
        return {
            "docKind": cost["docKind"],
            "docKindLabel": cost["docKindLabel"],
            "docId": cost["docId"],
            "docNo": cost["docNo"],
            "counterparty": cost["payeeName"],
            "amount": float(allocated if allocated is not None else cost["amount"]),
            "docAmount": float(cost["amount"]),
            "amountExclTax": float(cost["amountExclTax"]),
            "periodEnd": cost["periodEnd"],
        }

    @staticmethod
    def _assert_tax_mode(tax_mode: str) -> None:
        if tax_mode not in TAX_MODES:
            raise BizException("税额口径只能选含税或不含税")


def _money(v: Any) -> Decimal:
    return Decimal(str(v)).quantize(_CENT, rounding=ROUND_HALF_UP)


def _excl(amount: Decimal, rate: Decimal) -> Decimal:
    """含税转不含税：``amount ÷ (1 + 税率)``。"""
    r = Decimal(str(rate or 0))
    if r <= 0:
        return _money(amount)
    return _money(Decimal(str(amount)) / (1 + r / 100))


def _day_start(day: ddate) -> datetime:
    return datetime.combine(day, datetime.min.time())


def _day_end(day: ddate) -> datetime:
    return datetime.combine(day, datetime.max.time())
