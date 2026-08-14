"""应收账龄 Service（文档 12 §二、§五）

账龄不建表，全部由「客户结算单 + 收款核销」实时聚合：结算单是「确认了要收多少
钱」的单据，故账龄的原子单位是结算单，未收余额 =
``planned_amount - received_amount_total``。

到期日也不落库（结算单 ``due_date`` 只作单据级特批覆盖）：账期天数是客户档案上
的可变配置，存死值就要在改配置时批量刷历史数据，改成每次按客户档案推导。
"""

import json
import logging
from dataclasses import dataclass
from datetime import date as ddate, datetime, timedelta
from decimal import Decimal
from io import BytesIO
from typing import Any, Dict, List, Optional, Sequence, Tuple

from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.client.models.finance.customer_recon import CustomerRecon
from app.modules.client.models.finance.customer_settlement import (
    CustomerSettleReconLink,
    CustomerSettlement,
)
from app.modules.client.models.partner.customer import Customer
from app.modules.client.services.finance.base.constants import (
    CreditStatus,
    SettlementType,
)
from app.modules.client.services.finance.base.finance_state_machine import (
    FIN_PAID,
    FIN_REVIEWED,
)

logger = logging.getLogger(__name__)

# ``biz_system_config`` 约定：分桶阈值按租户可配，账期习惯差异大
AGING_BUCKETS_CONFIG_KEY = "finance.aging_buckets"
AGING_BUCKETS_CONFIG_GROUP = "finance"
AGING_BUCKETS_DESCRIPTION = (
    "应收账龄分桶阈值 JSON 数组（逾期天数上界，升序），默认 [30,60,90]，"
    "即「未到期 / 1-30 / 31-60 / 61-90 / 90 天以上」五桶"
)

DEFAULT_AGING_BUCKETS: Tuple[int, ...] = (30, 60, 90)
MAX_BUCKETS = 6

# 分转对齐误差：余额小于半分即视为收平
BALANCE_TOLERANCE = Decimal("0.005")


def default_aging_buckets_json() -> str:
    """默认配置 JSON（供开户种子 / 懒补齐使用）。"""
    return json.dumps(list(DEFAULT_AGING_BUCKETS), separators=(",", ":"))


def parse_aging_buckets(config_value: Optional[str]) -> List[int]:
    """解析分桶阈值，非法配置一律回退默认，不让报表因配错而打不开。"""
    if not config_value:
        return list(DEFAULT_AGING_BUCKETS)
    try:
        obj = json.loads(config_value)
        raw = obj.get("buckets") if isinstance(obj, dict) else obj
        values = sorted({int(x) for x in raw if int(x) > 0})
    except (AttributeError, TypeError, ValueError):
        logger.warning("账龄分桶阈值解析失败，回退默认: %r", config_value)
        return list(DEFAULT_AGING_BUCKETS)
    if not values or len(values) > MAX_BUCKETS:
        logger.warning("账龄分桶阈值数量不合法，回退默认: %r", config_value)
        return list(DEFAULT_AGING_BUCKETS)
    return values


def bucket_labels(buckets: Sequence[int]) -> List[str]:
    """桶名：未到期 + 各区间 + 最后一档以上。"""
    labels = ["未到期"]
    low = 1
    for upper in buckets:
        labels.append(f"{low}-{upper} 天")
        low = upper + 1
    labels.append(f"{buckets[-1]} 天以上")
    return labels


def bucket_index(overdue_days: int, buckets: Sequence[int]) -> int:
    """逾期天数落在第几桶（0 = 未到期）。"""
    if overdue_days <= 0:
        return 0
    for i, upper in enumerate(buckets):
        if overdue_days <= upper:
            return i + 1
    return len(buckets) + 1


@dataclass
class AgingRow:
    """一张有未收余额的结算单在账龄视图里的样子"""

    settle_id: int
    doc_no: str
    customer_id: int
    customer_name: Optional[str]
    enterprise_id: Optional[int]
    status: int
    planned: Decimal
    received: Decimal
    unpaid: Decimal
    due_date: Optional[ddate]
    overdue_days: int
    bucket: int
    period_start: Optional[ddate]
    period_end: Optional[ddate]
    settlement_type: Optional[int]
    payment_days: Optional[int]
    credit_limit: Optional[Decimal]
    credit_status: int
    due_date_overridden: bool

    def to_detail(self, labels: Sequence[str]) -> dict:
        return {
            "settleId": self.settle_id,
            "docNo": self.doc_no,
            "customerId": self.customer_id,
            "customerName": self.customer_name,
            "enterpriseId": self.enterprise_id,
            "status": self.status,
            "plannedAmount": float(self.planned),
            "receivedAmount": float(self.received),
            "unpaidAmount": float(self.unpaid),
            "dueDate": self.due_date,
            "dueDateOverridden": self.due_date_overridden,
            "overdueDays": self.overdue_days,
            "bucket": self.bucket,
            "bucketLabel": labels[self.bucket],
            "periodStart": self.period_start,
            "periodEnd": self.period_end,
        }


class AgingService:
    """应收账龄聚合"""

    # ------------------------------------------------------------------
    # 配置
    # ------------------------------------------------------------------
    @staticmethod
    async def load_buckets(db: AsyncSession) -> List[int]:
        """读取租户分桶阈值。"""
        from app.modules.client.services.system_config_service import (
            SystemConfigService,
        )

        raw = await SystemConfigService.get_by_key(db, AGING_BUCKETS_CONFIG_KEY)
        return parse_aging_buckets(raw)

    # ------------------------------------------------------------------
    # 明细行
    # ------------------------------------------------------------------
    @classmethod
    async def list_rows(
        cls,
        db: AsyncSession,
        *,
        base_date: Optional[ddate] = None,
        customer_id: Optional[int] = None,
        enterprise_id: Optional[int] = None,
        credit_status: Optional[int] = None,
        keyword: Optional[str] = None,
        buckets: Optional[Sequence[int]] = None,
    ) -> Tuple[List[AgingRow], List[int]]:
        """全部有未收余额的结算单（一次查询，到期日与分桶在内存算）。"""
        thresholds = list(buckets) if buckets else await cls.load_buckets(db)
        base = base_date or ddate.today()

        period = (
            select(
                CustomerSettleReconLink.settle_id.label("settle_id"),
                func.min(CustomerRecon.period_start).label("period_start"),
                func.max(CustomerRecon.period_end).label("period_end"),
            )
            .join(
                CustomerRecon,
                CustomerRecon.id == CustomerSettleReconLink.recon_id,
            )
            .where(
                CustomerSettleReconLink.is_deleted == 0,
                CustomerRecon.is_deleted == 0,
            )
            .group_by(CustomerSettleReconLink.settle_id)
            .subquery()
        )

        stmt = (
            select(
                CustomerSettlement,
                Customer.settlement_type,
                Customer.payment_days,
                Customer.credit_limit,
                Customer.credit_status,
                period.c.period_start,
                period.c.period_end,
            )
            .outerjoin(Customer, Customer.id == CustomerSettlement.customer_id)
            .outerjoin(period, period.c.settle_id == CustomerSettlement.id)
            .where(
                CustomerSettlement.is_deleted == 0,
                # 未审批的不算确认的应收；已收款态靠余额自然过滤（文档 12 §2.1）
                CustomerSettlement.status.in_([FIN_REVIEWED, FIN_PAID]),
                CustomerSettlement.planned_amount
                > CustomerSettlement.received_amount_total,
            )
        )
        if customer_id:
            stmt = stmt.where(CustomerSettlement.customer_id == customer_id)
        if enterprise_id:
            stmt = stmt.where(CustomerSettlement.enterprise_id == enterprise_id)
        if credit_status is not None:
            stmt = stmt.where(Customer.credit_status == credit_status)
        if keyword:
            kw = f"%{keyword.strip()}%"
            stmt = stmt.where(
                CustomerSettlement.customer_name.like(kw)
                | CustomerSettlement.doc_no.like(kw)
            )

        result = await db.execute(stmt.order_by(CustomerSettlement.id.desc()))
        rows: List[AgingRow] = []
        for (
            settle, s_type, pay_days, credit_limit, cs, p_start, p_end,
        ) in result.all():
            planned = Decimal(str(settle.planned_amount or 0))
            received = Decimal(str(settle.received_amount_total or 0))
            unpaid = planned - received
            if unpaid <= BALANCE_TOLERANCE:
                continue
            due = cls.resolve_due_date(
                settle,
                settlement_type=s_type,
                payment_days=pay_days,
                period_end=p_end,
            )
            overdue = (base - due).days if due else 0
            rows.append(AgingRow(
                settle_id=int(settle.id),
                doc_no=settle.doc_no,
                customer_id=int(settle.customer_id),
                customer_name=settle.customer_name,
                enterprise_id=settle.enterprise_id,
                status=int(settle.status),
                planned=planned,
                received=received,
                unpaid=unpaid,
                due_date=due,
                overdue_days=overdue,
                bucket=bucket_index(overdue, thresholds),
                period_start=_as_date(p_start),
                period_end=_as_date(p_end),
                settlement_type=s_type,
                payment_days=pay_days,
                credit_limit=(
                    Decimal(str(credit_limit)) if credit_limit is not None else None
                ),
                credit_status=int(cs if cs is not None else CreditStatus.NORMAL),
                due_date_overridden=settle.due_date is not None,
            ))
        return rows, thresholds

    @staticmethod
    def resolve_due_date(
        settle: Any,
        *,
        settlement_type: Optional[int],
        payment_days: Optional[int],
        period_end: Optional[datetime],
    ) -> Optional[ddate]:
        """到期日 = 账期起算日 + 账期天数（文档 12 §2.2）。

        单据上填了 ``due_date`` 就直接用它——那是「这一单特批 90 天」的场景，
        不该被客户档案的通用账期覆盖掉。
        """
        if settle.due_date is not None:
            return _as_date(settle.due_date)

        days = int(payment_days or 0)
        approved = _as_date(settle.reviewed_at) or _as_date(settle.created_at)
        if settlement_type == SettlementType.MONTHLY:
            anchor = _as_date(period_end) or approved
        elif settlement_type == SettlementType.BY_INVOICE:
            # 票到才起算；销项发票表第 4 期建，暂以审批日兜底
            anchor = approved
        elif settlement_type == SettlementType.PREPAY:
            # 预付客户理论上不该有应收，出现即视为异常，账期按 0 天
            anchor, days = approved, 0
        else:
            anchor = approved
        if anchor is None:
            return None
        return anchor + timedelta(days=days)

    # ------------------------------------------------------------------
    # 客户汇总
    # ------------------------------------------------------------------
    @classmethod
    async def customer_page(
        cls,
        db: AsyncSession,
        *,
        page: int = 1,
        page_size: int = 20,
        base_date: Optional[ddate] = None,
        customer_id: Optional[int] = None,
        enterprise_id: Optional[int] = None,
        credit_status: Optional[int] = None,
        keyword: Optional[str] = None,
        bucket: Optional[int] = None,
        only_overdue: bool = False,
        only_exceeded: bool = False,
    ) -> dict:
        """客户维度汇总（催收主视图，按未收余额倒序）。"""
        rows, thresholds = await cls.list_rows(
            db,
            base_date=base_date, customer_id=customer_id,
            enterprise_id=enterprise_id, credit_status=credit_status,
            keyword=keyword,
        )
        labels = bucket_labels(thresholds)
        groups = cls._group_by_customer(rows, thresholds)

        # 桶下钻只影响展示金额，超额判定始终按该客户全部未收余额
        if bucket is not None and bucket < len(labels):
            groups = [
                g for g in groups if g["bucketSummary"][bucket]["amount"] > 0
            ]
        if only_overdue:
            groups = [g for g in groups if g["overdueAmount"] > 0]
        if only_exceeded:
            groups = [g for g in groups if g["exceeded"]]

        total = len(groups)
        start = (max(1, page) - 1) * page_size
        return {
            "list": groups[start:start + page_size],
            "count": total,
            "total": total,
            "page": page,
            "page_size": page_size,
            "baseDate": base_date or ddate.today(),
            "buckets": thresholds,
            "bucketLabels": labels,
        }

    @classmethod
    async def summary(
        cls,
        db: AsyncSession,
        *,
        base_date: Optional[ddate] = None,
        enterprise_id: Optional[int] = None,
        credit_status: Optional[int] = None,
        keyword: Optional[str] = None,
    ) -> dict:
        """KPI + 分桶分布 + 按经营主体（多法人必须分开看）。"""
        rows, thresholds = await cls.list_rows(
            db,
            base_date=base_date, enterprise_id=enterprise_id,
            credit_status=credit_status, keyword=keyword,
        )
        labels = bucket_labels(thresholds)
        groups = cls._group_by_customer(rows, thresholds)

        buckets_out = [
            {"bucket": i, "label": labels[i], "amount": 0.0, "count": 0}
            for i in range(len(labels))
        ]
        by_entity: Dict[Any, dict] = {}
        total_unpaid = Decimal("0")
        overdue_total = Decimal("0")
        for r in rows:
            buckets_out[r.bucket]["amount"] = round(
                buckets_out[r.bucket]["amount"] + float(r.unpaid), 2
            )
            buckets_out[r.bucket]["count"] += 1
            total_unpaid += r.unpaid
            if r.overdue_days > 0:
                overdue_total += r.unpaid
            slot = by_entity.setdefault(r.enterprise_id, {
                "enterpriseId": r.enterprise_id,
                "unpaidAmount": 0.0,
                "overdueAmount": 0.0,
            })
            slot["unpaidAmount"] = round(slot["unpaidAmount"] + float(r.unpaid), 2)
            if r.overdue_days > 0:
                slot["overdueAmount"] = round(
                    slot["overdueAmount"] + float(r.unpaid), 2
                )

        last_bucket_amount = buckets_out[-1]["amount"] if len(labels) > 1 else 0.0
        return {
            "baseDate": base_date or ddate.today(),
            "buckets": thresholds,
            "bucketLabels": labels,
            "kpi": {
                "totalUnpaid": float(total_unpaid),
                "notDueAmount": buckets_out[0]["amount"],
                "overdueAmount": float(overdue_total),
                "lastBucketLabel": labels[-1],
                "lastBucketAmount": last_bucket_amount,
                "customerCount": len(groups),
                "exceededCustomerCount": sum(1 for g in groups if g["exceeded"]),
                "settleCount": len(rows),
            },
            "bucketDistribution": buckets_out,
            "byEnterprise": sorted(
                by_entity.values(),
                key=lambda x: x["unpaidAmount"], reverse=True,
            ),
        }

    @classmethod
    async def customer_detail(
        cls,
        db: AsyncSession,
        customer_id: int,
        *,
        base_date: Optional[ddate] = None,
        bucket: Optional[int] = None,
    ) -> dict:
        """某客户的结算单明细（客户汇总表展开用）。"""
        rows, thresholds = await cls.list_rows(
            db, base_date=base_date, customer_id=customer_id,
        )
        labels = bucket_labels(thresholds)
        shown = [r for r in rows if bucket is None or r.bucket == bucket]
        shown.sort(key=lambda r: (r.due_date or ddate.max, -float(r.unpaid)))
        groups = cls._group_by_customer(rows, thresholds)
        return {
            "baseDate": base_date or ddate.today(),
            "buckets": thresholds,
            "bucketLabels": labels,
            "customer": groups[0] if groups else None,
            "list": [r.to_detail(labels) for r in shown],
            "count": len(shown),
        }

    # ------------------------------------------------------------------
    # 单客户摘要（业务侧页面同步调用，要快）
    # ------------------------------------------------------------------
    @classmethod
    async def customer_brief(
        cls,
        db: AsyncSession,
        customer_id: int,
        *,
        base_date: Optional[ddate] = None,
    ) -> dict:
        """未收余额 / 逾期 / 额度 / 超额 一次给全。

        信用字段优先取明细行上的冗余值，只有该客户完全没有未收余额时才回查客户
        档案——正常路径两次查询（配置 + 明细），不做复杂 join。
        """
        rows, thresholds = await cls.list_rows(
            db, base_date=base_date, customer_id=customer_id,
        )
        labels = bucket_labels(thresholds)
        if rows:
            brief = cls._group_by_customer(rows, thresholds)[0]
            brief["bucketLabels"] = labels
            brief["buckets"] = thresholds
            return brief

        customer = (await db.execute(
            select(Customer).where(
                Customer.id == customer_id, Customer.is_deleted == 0,
            )
        )).scalar_one_or_none()
        limit = (
            Decimal(str(customer.credit_limit))
            if customer is not None and customer.credit_limit is not None
            else None
        )
        return {
            "customerId": customer_id,
            "customerName": customer.customer_name if customer else None,
            "enterpriseId": customer.enterprise_id if customer else None,
            "creditStatus": (
                int(customer.credit_status) if customer else CreditStatus.NORMAL
            ),
            "creditStatusLabel": CreditStatus.LABELS.get(
                int(customer.credit_status) if customer else CreditStatus.NORMAL,
                "正常",
            ),
            "creditLimit": float(limit) if limit is not None else None,
            "unpaidAmount": 0.0,
            "overdueAmount": 0.0,
            "maxOverdueDays": 0,
            "settleCount": 0,
            "exceeded": False,
            "exceededAmount": 0.0,
            "bucketSummary": [
                {"bucket": i, "label": labels[i], "amount": 0.0, "count": 0}
                for i in range(len(labels))
            ],
            "buckets": thresholds,
            "bucketLabels": labels,
        }

    # ------------------------------------------------------------------
    # 导出
    # ------------------------------------------------------------------
    @classmethod
    async def build_export_workbook(
        cls,
        db: AsyncSession,
        *,
        base_date: Optional[ddate] = None,
        enterprise_id: Optional[int] = None,
        credit_status: Optional[int] = None,
        keyword: Optional[str] = None,
        bucket: Optional[int] = None,
    ) -> bytes:
        """两张表页：客户汇总（催收用）+ 结算单明细（对账用）。

        表头带统计基准日与分桶阈值——不写口径的表格发出去，看的人无法判断
        「逾期 40 天」是按哪天算的。
        """
        try:
            import openpyxl  # type: ignore
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError as e:  # pragma: no cover
            from app.common.exceptions import BizException

            raise BizException("导出失败，服务端缺少表格组件，请联系管理员") from e

        rows, thresholds = await cls.list_rows(
            db,
            base_date=base_date, enterprise_id=enterprise_id,
            credit_status=credit_status, keyword=keyword,
        )
        labels = bucket_labels(thresholds)
        groups = cls._group_by_customer(rows, thresholds)
        shown = [r for r in rows if bucket is None or r.bucket == bucket]
        base = base_date or ddate.today()
        caption = (
            f"统计基准日 {base.isoformat()}；分桶阈值 "
            f"{'/'.join(str(x) for x in thresholds)} 天"
        )

        head_fill = PatternFill("solid", fgColor="FFE7EEF8")
        head_font = Font(bold=True, size=11)

        def _write(ws, headers: List[str], widths: Sequence[int]) -> None:
            ws.append([caption])
            ws.append(headers)
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=2, column=col)
                cell.fill = head_fill
                cell.font = head_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = w

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "客户汇总"
        head1 = ["客户", "信用状态", "信用额度", "未收余额", "逾期金额",
                 "最大逾期天数", "结算单数", "是否超额", "超额金额"] + labels
        _write(ws1, head1, [24, 12, 14, 14, 14, 14, 10, 10, 14] + [14] * len(labels))
        for g in groups:
            ws1.append([
                g["customerName"],
                g["creditStatusLabel"],
                g["creditLimit"],
                g["unpaidAmount"],
                g["overdueAmount"],
                g["maxOverdueDays"],
                g["settleCount"],
                "超额" if g["exceeded"] else "",
                g["exceededAmount"],
            ] + [c["amount"] for c in g["bucketSummary"]])

        ws2 = wb.create_sheet("结算单明细")
        head2 = ["结算单号", "客户", "对账周期起", "对账周期止", "应收金额",
                 "已收金额", "未收余额", "到期日", "逾期天数", "账龄桶"]
        _write(ws2, head2, [20, 24, 14, 14, 14, 14, 14, 14, 12, 14])
        for r in sorted(shown, key=lambda x: (x.due_date or ddate.max)):
            ws2.append([
                r.doc_no,
                r.customer_name,
                r.period_start,
                r.period_end,
                float(r.planned),
                float(r.received),
                float(r.unpaid),
                r.due_date,
                r.overdue_days,
                labels[r.bucket],
            ])

        bio = BytesIO()
        wb.save(bio)
        return bio.getvalue()

    # ------------------------------------------------------------------
    # 内部工具
    # ------------------------------------------------------------------
    @staticmethod
    def _group_by_customer(
        rows: Sequence[AgingRow], thresholds: Sequence[int],
    ) -> List[dict]:
        """按客户汇总，未收余额倒序。"""
        labels = bucket_labels(thresholds)
        acc: Dict[int, dict] = {}
        for r in rows:
            slot = acc.get(r.customer_id)
            if slot is None:
                slot = acc[r.customer_id] = {
                    "customerId": r.customer_id,
                    "customerName": r.customer_name,
                    "enterpriseId": r.enterprise_id,
                    "creditStatus": r.credit_status,
                    "creditStatusLabel": CreditStatus.LABELS.get(
                        r.credit_status, "正常"
                    ),
                    "creditLimit": (
                        float(r.credit_limit) if r.credit_limit is not None else None
                    ),
                    "unpaidAmount": Decimal("0"),
                    "overdueAmount": Decimal("0"),
                    "maxOverdueDays": 0,
                    "settleCount": 0,
                    "bucketSummary": [
                        {"bucket": i, "label": labels[i],
                         "amount": Decimal("0"), "count": 0}
                        for i in range(len(labels))
                    ],
                }
            slot["unpaidAmount"] += r.unpaid
            slot["settleCount"] += 1
            if r.overdue_days > 0:
                slot["overdueAmount"] += r.unpaid
                slot["maxOverdueDays"] = max(
                    int(slot["maxOverdueDays"]), r.overdue_days
                )
            cell = slot["bucketSummary"][r.bucket]
            cell["amount"] += r.unpaid
            cell["count"] += 1

        out: List[dict] = []
        for slot in acc.values():
            unpaid = slot["unpaidAmount"]
            limit = slot["creditLimit"]
            exceeded_amount = (
                unpaid - Decimal(str(limit)) if limit is not None else Decimal("0")
            )
            slot["unpaidAmount"] = float(unpaid)
            slot["overdueAmount"] = float(slot["overdueAmount"])
            slot["exceeded"] = limit is not None and exceeded_amount > 0
            slot["exceededAmount"] = float(max(exceeded_amount, Decimal("0")))
            slot["bucketSummary"] = [
                {**c, "amount": float(c["amount"])} for c in slot["bucketSummary"]
            ]
            out.append(slot)
        out.sort(key=lambda x: x["unpaidAmount"], reverse=True)
        return out


def _as_date(value: Any) -> Optional[ddate]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, ddate):
        return value
    return None
